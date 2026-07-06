import json
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent


def as_decimal(value):
    return str(Decimal(str(value or 0)).quantize(Decimal("0.01")))


def inspect_workbook(filename):
    workbook = load_workbook(filename, data_only=True)
    result = {}
    for sheet_name in ("Ventes", "Achats"):
        worksheet = workbook[sheet_name]
        totals_row = None
        for row_idx in range(1, worksheet.max_row + 1):
            if worksheet.cell(row_idx, 1).value == "Nombre de documents":
                totals_row = row_idx
                break
        if totals_row is None:
            raise AssertionError(f"Totals row not found in {sheet_name}")

        amount_start = 6 if sheet_name == "Ventes" else 7
        result[sheet_name] = {
            "column_a_width": worksheet.column_dimensions["A"].width,
            "totals_row": totals_row,
            "label": worksheet.cell(totals_row, 1).value,
            "document_count": worksheet.cell(totals_row, 2).value,
            "amount_ht": as_decimal(worksheet.cell(totals_row, amount_start).value),
            "tva": as_decimal(worksheet.cell(totals_row, amount_start + 1).value),
            "amount_ttc": as_decimal(worksheet.cell(totals_row, amount_start + 2).value),
            "paid_or_sold": as_decimal(worksheet.cell(totals_row, amount_start + 3).value),
            "residual_or_balance": as_decimal(worksheet.cell(totals_row, amount_start + 4).value),
        }
    return result


def comparable_values(workbook_inspection):
    return {
        sheet: {
            key: value
            for key, value in values.items()
            if key not in {"column_a_width"}
        }
        for sheet, values in workbook_inspection.items()
    }


june = inspect_workbook(
    BASE_DIR / "REQA_R06_Rapport_facturation_La_Platine_2026-06-01_2026-06-30.xlsx"
)
empty = inspect_workbook(
    BASE_DIR / "REQA_R06_Rapport_facturation_La_Platine_2099-12-01_2099-12-31_EMPTY.xlsx"
)

june_reference = inspect_workbook(
    BASE_DIR.parent / "Rapport_facturation_La_Platine_2026-06-01_2026-06-30.xlsx"
)
empty_reference = inspect_workbook(
    BASE_DIR.parent / "Rapport_facturation_La_Platine_2099-12-01_2099-12-31_SMOKE_E_VIDE.xlsx"
)

inspection = {
    "june_2026": june,
    "empty_2099_12": empty,
    "unchanged_against_initial_qa": {
        "june_2026_counts_and_totals": comparable_values(june)
        == comparable_values(june_reference),
        "empty_2099_12_counts_and_totals": comparable_values(empty)
        == comparable_values(empty_reference),
    },
    "initial_qa_reference_column_a_widths": {
        "june_2026": {
            sheet: values["column_a_width"] for sheet, values in june_reference.items()
        },
        "empty_2099_12": {
            sheet: values["column_a_width"] for sheet, values in empty_reference.items()
        },
    },
    "re_qa_column_a_widths": {
        "june_2026": {
            sheet: values["column_a_width"] for sheet, values in june.items()
        },
        "empty_2099_12": {
            sheet: values["column_a_width"] for sheet, values in empty.items()
        },
    },
}

(BASE_DIR / "reqa_r06_inspection.json").write_text(
    json.dumps(inspection, indent=2, ensure_ascii=False) + "\n",
    encoding="utf-8",
)
print(json.dumps(inspection, indent=2, ensure_ascii=False))
