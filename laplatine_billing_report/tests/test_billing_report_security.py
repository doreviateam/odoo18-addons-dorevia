# -*- coding: utf-8 -*-
import base64
from io import BytesIO

from openpyxl import load_workbook
from odoo import fields
from odoo.exceptions import AccessError
from odoo.tests import TransactionCase, tagged


@tagged("post_install", "-at_install", "laplatine_billing_report")
class TestLaplatineBillingReportSecurity(TransactionCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Wizard = cls.env["laplatine.billing.report.wizard"]
        cls.invoice_group = cls.env.ref("account.group_account_invoice")
        cls.internal_group = cls.env.ref("base.group_user")
        cls.menu_billing = cls.env.ref(
            "laplatine_billing_report.menu_laplatine_billing_report"
        )
        cls.action_billing = cls.env.ref(
            "laplatine_billing_report.action_laplatine_billing_report_wizard"
        )
        cls.denied_user = cls.env["res.users"].create(
            {
                "name": "Billing Report Denied User",
                "login": "billing_report_denied_test",
                "groups_id": [(6, 0, [cls.internal_group.id])],
            }
        )
        cls.invoice_user = cls.env["res.users"].create(
            {
                "name": "Billing Report Invoice User",
                "login": "billing_report_invoice_test",
                "groups_id": [(6, 0, [cls.internal_group.id, cls.invoice_group.id])],
            }
        )

    def test_t22_unauthorized_user_cannot_create_wizard(self):
        """T22 — utilisateur sans groupe Facturation refusé."""
        wizard_env = self.Wizard.with_user(self.denied_user)
        with self.assertRaises(AccessError):
            wizard_env.create(
                {
                    "date_from": fields.Date.from_string("2099-06-01"),
                    "date_to": fields.Date.from_string("2099-06-30"),
                }
            )

    def test_t22_unauthorized_user_cannot_generate_or_download(self):
        """T22 — pas de génération ni lecture binaire sans droit."""
        wizard = self.Wizard.create(
            {
                "date_from": fields.Date.from_string("2099-06-01"),
                "date_to": fields.Date.from_string("2099-06-30"),
            }
        )
        denied_wizard = wizard.with_user(self.denied_user)
        with self.assertRaises(AccessError):
            denied_wizard.action_generate_xlsx()
        wizard.action_generate_xlsx()
        with self.assertRaises(AccessError):
            denied_wizard.read(["report_file", "report_filename"])

    def test_t22_no_ir_attachment_created_on_generate(self):
        """T22 — aucun ir.attachment permanent créé."""
        attachment_count_before = self.env["ir.attachment"].search_count([])
        wizard = self.Wizard.with_user(self.invoice_user).create(
            {
                "date_from": fields.Date.from_string("2099-12-01"),
                "date_to": fields.Date.from_string("2099-12-31"),
            }
        )
        wizard.action_generate_xlsx()
        attachment_count_after = self.env["ir.attachment"].search_count([])
        self.assertEqual(attachment_count_before, attachment_count_after)

    def test_e05_invoice_user_can_access_menu_and_action(self):
        """Slice E — menu visible et droits wizard pour le groupe Facturation."""
        menu_env = self.env["ir.ui.menu"].with_user(self.invoice_user)
        self.assertTrue(menu_env.search([("id", "=", self.menu_billing.id)]))

        action = self.action_billing.sudo()
        self.assertIn(self.invoice_group, action.groups_id)
        self.assertEqual(action.res_model, "laplatine.billing.report.wizard")

        wizard_env = self.Wizard.with_user(self.invoice_user)
        self.assertTrue(wizard_env.check_access_rights("create", raise_exception=False))
        self.assertTrue(wizard_env.check_access_rights("read", raise_exception=False))

    def test_e06_denied_user_cannot_see_menu(self):
        """Slice E — menu masqué sans groupe Facturation."""
        menu_env = self.env["ir.ui.menu"].with_user(self.denied_user)
        self.assertFalse(menu_env.search([("id", "=", self.menu_billing.id)]))

    def test_e07_invoice_user_can_generate_report(self):
        """Slice E — profil autorisé génère le fichier."""
        wizard = self.Wizard.with_user(self.invoice_user).create(
            {
                "date_from": fields.Date.from_string("2099-12-01"),
                "date_to": fields.Date.from_string("2099-12-31"),
            }
        )
        action = wizard.action_generate_xlsx()
        self.assertEqual(action["type"], "ir.actions.act_url")
        self.assertTrue(wizard.report_file)

        workbook = load_workbook(BytesIO(base64.b64decode(wizard.report_file)))
        self.assertEqual(workbook.sheetnames, ["Ventes", "Achats"])

    def test_e08_laplatine_menu_between_vendors_and_accounting(self):
        """Menu La Platine entre Fournisseurs et Comptabilité sous Facturation."""
        parent = self.env.ref("account.menu_finance")
        vendors = self.env.ref("account.menu_finance_payables")
        accounting = self.env.ref("account.menu_finance_entries")
        laplatine = self.menu_billing.parent_id

        siblings = self.env["ir.ui.menu"].search(
            [("parent_id", "=", parent.id)], order="sequence, id"
        )
        sibling_ids = siblings.ids
        vendors_idx = sibling_ids.index(vendors.id)
        laplatine_idx = sibling_ids.index(laplatine.id)
        accounting_idx = sibling_ids.index(accounting.id)

        self.assertLess(vendors_idx, laplatine_idx)
        self.assertLess(laplatine_idx, accounting_idx)
        self.assertGreaterEqual(laplatine.sequence, vendors.sequence)
        self.assertLess(laplatine.sequence, accounting.sequence)
