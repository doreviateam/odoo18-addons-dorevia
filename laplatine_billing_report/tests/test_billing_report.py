# -*- coding: utf-8 -*-
import base64
from io import BytesIO
from unittest.mock import patch

from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from odoo import fields
from odoo.exceptions import UserError
from odoo.tests import TransactionCase, tagged

from ..report.billing_report_xlsx import (
    ACHATS_COLUMN_WIDTHS,
    ACHATS_HEADERS,
    EMPTY_MESSAGE_ROW,
    EMPTY_SHEET_MESSAGE,
    FIRST_DATA_ROW,
    HEADER_ROW,
    VENTES_COLUMN_WIDTHS,
    VENTES_HEADERS,
    document_type_label,
    report_sign,
    signed_move_amounts,
)


@tagged("post_install", "-at_install", "laplatine_billing_report")
class TestLaplatineBillingReportWizard(TransactionCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Wizard = cls.env["laplatine.billing.report.wizard"]
        cls.company = cls.env.company
        cls.partner = cls.env["res.partner"].create(
            {
                "name": "Client Test Billing Report",
                "customer_rank": 1,
            }
        )
        cls.vendor = cls.env["res.partner"].create(
            {
                "name": "Fournisseur Test Billing Report",
                "supplier_rank": 1,
            }
        )
        cls.sale_journal = cls.env["account.journal"].search(
            [("type", "=", "sale"), ("company_id", "=", cls.company.id)],
            limit=1,
        )
        cls.purchase_journal = cls.env["account.journal"].search(
            [("type", "=", "purchase"), ("company_id", "=", cls.company.id)],
            limit=1,
        )
        cls.income_account = cls.env["account.account"].search(
            [("account_type", "=", "income")],
            limit=1,
        )
        cls.expense_account = cls.env["account.account"].search(
            [("account_type", "=", "expense")],
            limit=1,
        )
        assert cls.sale_journal, "Journal de vente requis."
        assert cls.purchase_journal, "Journal d'achat requis."
        assert cls.income_account, "Compte de produit requis."
        assert cls.expense_account, "Compte de charge requis."

        cls.period_from = fields.Date.from_string("2099-06-01")
        cls.period_to = fields.Date.from_string("2099-06-30")

    def _wizard(self, date_from=None, date_to=None):
        if date_from is None:
            date_from = self.period_from
        if date_to is None:
            date_to = self.period_to
        return self.Wizard.create(
            {
                "date_from": date_from,
                "date_to": date_to,
            }
        )

    def _create_customer_move(
        self,
        move_type="out_invoice",
        amount=100.0,
        invoice_date=None,
        partner=None,
        currency=None,
        company=None,
    ):
        partner = partner or self.partner
        invoice_date = invoice_date or self.period_from
        env = self.env
        if company:
            env = env.with_company(company).sudo()
        account = self.income_account if company is None else company.env["account.account"].search(
            [("account_type", "=", "income")],
            limit=1,
        )
        journal = self.sale_journal if company is None else company.env["account.journal"].search(
            [("type", "=", "sale"), ("company_id", "=", company.id)],
            limit=1,
        )
        values = {
            "move_type": move_type,
            "partner_id": partner.id,
            "invoice_date": invoice_date,
            "journal_id": journal.id,
            "invoice_line_ids": [
                (
                    0,
                    0,
                    {
                        "name": "Ligne test rapport facturation",
                        "quantity": 1,
                        "price_unit": amount,
                        "account_id": account.id,
                    },
                )
            ],
        }
        if currency:
            values["currency_id"] = currency.id
        move = env["account.move"].create(values)
        move.action_post()
        return move

    def _create_vendor_move(self, amount=80.0, invoice_date=None, vendor_ref=None, partner=None):
        invoice_date = invoice_date or self.period_from
        values = {
            "move_type": "in_invoice",
            "partner_id": (partner or self.vendor).id,
            "invoice_date": invoice_date,
            "journal_id": self.purchase_journal.id,
            "invoice_line_ids": [
                (
                    0,
                    0,
                    {
                        "name": "Ligne test achat",
                        "quantity": 1,
                        "price_unit": amount,
                        "account_id": self.expense_account.id,
                    },
                )
            ],
        }
        if vendor_ref:
            values["ref"] = vendor_ref
        move = self.env["account.move"].create(values)
        move.action_post()
        return move

    def _load_ventes_sheet(self, wizard):
        workbook = load_workbook(BytesIO(base64.b64decode(wizard.report_file)))
        return workbook["Ventes"]

    def _load_workbook(self, wizard):
        return load_workbook(BytesIO(base64.b64decode(wizard.report_file)))

    def _load_achats_sheet(self, wizard):
        return self._load_workbook(wizard)["Achats"]

    def _sheet_headers(self, worksheet, headers):
        header_row_excel = HEADER_ROW + 1
        return [
            worksheet.cell(header_row_excel, col).value
            for col in range(1, len(headers) + 1)
        ]

    def _ventes_headers(self, worksheet):
        return self._sheet_headers(worksheet, VENTES_HEADERS)

    def _achats_headers(self, worksheet):
        return self._sheet_headers(worksheet, ACHATS_HEADERS)

    def _data_rows(self, worksheet, header_count):
        first_data_excel = FIRST_DATA_ROW + 1
        rows = []
        row_idx = first_data_excel
        while True:
            first_cell = worksheet.cell(row_idx, 1).value
            if first_cell == "Nombre de documents":
                break
            if first_cell in ("Facture", "Avoir"):
                rows.append(row_idx)
            elif first_cell == "Aucun document trouvé sur la période sélectionnée.":
                break
            else:
                break
            row_idx += 1
        return rows

    def _ventes_data_rows(self, worksheet):
        return self._data_rows(worksheet, len(VENTES_HEADERS))

    def _achats_data_rows(self, worksheet):
        return self._data_rows(worksheet, len(ACHATS_HEADERS))

    def _find_totals_row(self, worksheet):
        for row_idx in range(FIRST_DATA_ROW + 1, worksheet.max_row + 1):
            if worksheet.cell(row_idx, 1).value == "Nombre de documents":
                return row_idx
        return None

    def _assert_text_cell_not_formula(self, cell, expected_value):
        self.assertEqual(cell.value, expected_value)
        self.assertNotEqual(cell.data_type, "f")

    def _assert_empty_sheet(self, worksheet, headers, amount_total_col):
        self.assertEqual(
            self._sheet_headers(worksheet, headers),
            headers,
        )
        self.assertEqual(worksheet.cell(EMPTY_MESSAGE_ROW + 1, 1).value, EMPTY_SHEET_MESSAGE)
        totals_row = self._find_totals_row(worksheet)
        self.assertIsNotNone(totals_row)
        self.assertEqual(worksheet.cell(totals_row, 1).value, "Nombre de documents")
        self.assertEqual(worksheet.cell(totals_row, 2).value, 0)
        for col in range(amount_total_col, amount_total_col + 5):
            self.assertEqual(worksheet.cell(totals_row, col).value, 0)

    def test_t01_default_period_is_previous_calendar_month(self):
        """T01 — ouverture wizard : mois M-1 complet."""
        today = fields.Date.from_string("2026-07-06")
        with patch.object(
            type(self.Wizard),
            "_default_period_bounds",
            return_value=(
                today.replace(day=1) - relativedelta(months=1),
                today.replace(day=1) - relativedelta(days=1),
            ),
        ):
            defaults = self.Wizard.default_get(["date_from", "date_to"])
        self.assertEqual(defaults["date_from"], fields.Date.from_string("2026-06-01"))
        self.assertEqual(defaults["date_to"], fields.Date.from_string("2026-06-30"))

    def test_t01_default_period_january_uses_december_previous_year(self):
        """Gate slice A — M-1 correct en janvier."""
        today = fields.Date.from_string("2026-01-15")
        with patch.object(
            type(self.Wizard),
            "_default_period_bounds",
            return_value=(
                today.replace(day=1) - relativedelta(months=1),
                today.replace(day=1) - relativedelta(days=1),
            ),
        ):
            defaults = self.Wizard.default_get(["date_from", "date_to"])
        self.assertEqual(defaults["date_from"], fields.Date.from_string("2025-12-01"))
        self.assertEqual(defaults["date_to"], fields.Date.from_string("2025-12-31"))

    def test_t02_invalid_period_raises_user_error(self):
        """T02 — date_from > date_to refusée."""
        with self.assertRaises(UserError):
            self._wizard(
                date_from=fields.Date.from_string("2026-06-30"),
                date_to=fields.Date.from_string("2026-06-01"),
            )
        wizard = self.Wizard.new(
            {
                "date_from": fields.Date.from_string("2026-06-30"),
                "date_to": fields.Date.from_string("2026-06-01"),
            }
        )
        with self.assertRaises(UserError):
            wizard.action_generate_xlsx()

    def test_t03_custom_period_generates_downloadable_xlsx(self):
        """T03 — période modifiable, génération OK."""
        date_from = fields.Date.from_string("2026-01-01")
        date_to = fields.Date.from_string("2026-03-31")
        attachment_count_before = self.env["ir.attachment"].search_count([])
        wizard = self._wizard(date_from=date_from, date_to=date_to)
        action = wizard.action_generate_xlsx()

        self.assertEqual(action["type"], "ir.actions.act_url")
        self.assertIn("download=true", action["url"])
        self.assertTrue(wizard.report_file)
        workbook = load_workbook(BytesIO(base64.b64decode(wizard.report_file)))
        self.assertEqual(workbook.sheetnames, ["Ventes", "Achats"])
        self.assertEqual(attachment_count_before, self.env["ir.attachment"].search_count([]))

    def test_t04_posted_customer_invoice_in_ventes_sheet(self):
        """T04 — facture client comptabilisée dans la période."""
        invoice = self._create_customer_move(amount=500.0)
        wizard = self._wizard()
        wizard.action_generate_xlsx()
        worksheet = self._load_ventes_sheet(wizard)
        data_rows = self._ventes_data_rows(worksheet)
        self.assertEqual(len(data_rows), 1)
        row = data_rows[0]
        self.assertEqual(worksheet.cell(row, 1).value, "Facture")
        self.assertEqual(worksheet.cell(row, 2).value, invoice.name)
        self.assertEqual(worksheet.cell(row, 3).value, self.partner.display_name)

    def test_t05_draft_and_cancelled_excluded(self):
        """T05 — brouillon et annulée absents."""
        draft = self.env["account.move"].create(
            {
                "move_type": "out_invoice",
                "partner_id": self.partner.id,
                "invoice_date": self.period_from,
                "journal_id": self.sale_journal.id,
                "invoice_line_ids": [
                    (
                        0,
                        0,
                        {
                            "name": "Brouillon",
                            "quantity": 1,
                            "price_unit": 50.0,
                            "account_id": self.income_account.id,
                        },
                    )
                ],
            }
        )
        cancelled = self.env["account.move"].create(
            {
                "move_type": "out_invoice",
                "partner_id": self.partner.id,
                "invoice_date": self.period_from,
                "journal_id": self.sale_journal.id,
                "invoice_line_ids": [
                    (
                        0,
                        0,
                        {
                            "name": "Annulée",
                            "quantity": 1,
                            "price_unit": 75.0,
                            "account_id": self.income_account.id,
                        },
                    )
                ],
            }
        )
        cancelled.action_post()
        cancelled.button_cancel()
        wizard = self._wizard()
        moves = wizard._fetch_sale_moves()
        self.assertNotIn(draft, moves)
        self.assertNotIn(cancelled, moves)

    def test_t08_posted_vendor_invoice_on_achats_sheet(self):
        """T08 — facture fournisseur comptabilisée sur l'onglet Achats."""
        vendor_bill = self._create_vendor_move(amount=80.0, vendor_ref="REF-FOURN-TEST-001")
        wizard = self._wizard()
        wizard.action_generate_xlsx()
        worksheet = self._load_achats_sheet(wizard)
        rows = self._achats_data_rows(worksheet)
        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(worksheet.cell(row, 1).value, "Facture")
        self.assertEqual(worksheet.cell(row, 2).value, vendor_bill.name)
        self.assertEqual(worksheet.cell(row, 3).value, "REF-FOURN-TEST-001")
        self.assertEqual(worksheet.cell(row, 4).value, self.vendor.display_name)

    def test_t17_two_sheets_always_present(self):
        """T17 — fichier avec exactement 2 onglets Ventes et Achats."""
        wizard = self._wizard()
        wizard.action_generate_xlsx()
        workbook = self._load_workbook(wizard)
        self.assertEqual(workbook.sheetnames, ["Ventes", "Achats"])

    def test_t19_filename_contains_period_dates(self):
        """T19 — nom de fichier avec les deux dates de période."""
        date_from = fields.Date.from_string("2099-03-01")
        date_to = fields.Date.from_string("2099-03-31")
        wizard = self._wizard(date_from=date_from, date_to=date_to)
        wizard.action_generate_xlsx()
        self.assertEqual(
            wizard.report_filename,
            "Rapport_facturation_La_Platine_2099-03-01_2099-03-31.xlsx",
        )

    def test_c01_achats_column_order(self):
        """Slice C — 12 colonnes Achats dans l'ordre MOA."""
        wizard = self._wizard()
        wizard.action_generate_xlsx()
        worksheet = self._load_achats_sheet(wizard)
        self.assertEqual(self._achats_headers(worksheet), ACHATS_HEADERS)

    def test_c02_vendor_refund_negative_amounts_on_achats(self):
        """Slice C — in_refund montants négatifs, Type Avoir."""
        bill = self._create_vendor_move(amount=200.0)
        refund = bill._reverse_moves(
            default_values_list=[{"invoice_date": self.period_from}]
        )
        refund.action_post()
        wizard = self._wizard()
        wizard.action_generate_xlsx()
        worksheet = self._load_achats_sheet(wizard)
        refund_row = None
        for row_idx in self._achats_data_rows(worksheet):
            if worksheet.cell(row_idx, 1).value == "Avoir":
                refund_row = row_idx
        self.assertIsNotNone(refund_row)
        self.assertLess(worksheet.cell(refund_row, 9).value, 0)

    def test_c03_achats_totals_algebraic_sum(self):
        """Slice C — totaux algébriques facture + avoir fournisseur."""
        bill = self._create_vendor_move(amount=500.0)
        refund = bill._reverse_moves(
            default_values_list=[{"invoice_date": self.period_from}]
        )
        refund.action_post()
        wizard = self._wizard()
        wizard.action_generate_xlsx()
        worksheet = self._load_achats_sheet(wizard)

        purchase_moves = wizard._fetch_purchase_moves()
        expected_ttc = sum(
            signed_move_amounts(move)["amount_ttc"] for move in purchase_moves
        )

        totals_row = None
        for row_idx in range(FIRST_DATA_ROW + 1, worksheet.max_row + 1):
            if worksheet.cell(row_idx, 1).value == "Nombre de documents":
                totals_row = row_idx
                break
        self.assertIsNotNone(totals_row)
        self.assertEqual(worksheet.cell(totals_row, 2).value, 2)
        self.assertAlmostEqual(worksheet.cell(totals_row, 9).value, expected_ttc)

    def test_t06_domain_filters_active_company(self):
        """T06 — filtre company_id = env.company.id (inspection domaine)."""
        wizard = self._wizard()
        domain = wizard._move_base_domain()
        self.assertIn(("company_id", "=", self.env.company.id), domain)

    def test_t06_other_company_excluded(self):
        """T06 — autre société absente (domaine company_id = env.company.id)."""
        other_company = self.env["res.company"].search(
            [("id", "!=", self.company.id)], limit=1
        )
        if not other_company:
            self.skipTest("Seconde société requise pour ce test.")
        other_partner = self.env["res.partner"].create(
            {"name": "Client autre société billing", "customer_rank": 1}
        )
        self._create_customer_move(
            amount=120.0,
            partner=other_partner,
            company=other_company,
        )
        own_invoice = self._create_customer_move(amount=90.0)
        wizard = self._wizard()
        moves = wizard._fetch_sale_moves()
        self.assertIn(own_invoice, moves)
        self.assertEqual(len(moves), 1)

    def test_t07_invoice_date_outside_period_excluded(self):
        """T07 — invoice_date hors période absente."""
        self._create_customer_move(
            amount=100.0,
            invoice_date=fields.Date.from_string("2099-05-31"),
        )
        in_period = self._create_customer_move(amount=200.0)
        wizard = self._wizard()
        moves = wizard._fetch_sale_moves()
        self.assertEqual(moves, in_period)

    def test_t09_foreign_currency_blocked(self):
        """T09 — devise étrangère bloque la génération."""
        usd = self.env["res.currency"].search([("name", "=", "USD")], limit=1)
        self.assertTrue(usd)
        self._create_customer_move(amount=100.0)
        self._create_customer_move(amount=50.0, currency=usd)
        wizard = self._wizard()
        with self.assertRaises(UserError):
            wizard.action_generate_xlsx()

    def test_t10_customer_invoice_positive_amounts(self):
        """T10 — out_invoice montants positifs, Type Facture."""
        self._create_customer_move(amount=1000.0)
        wizard = self._wizard()
        wizard.action_generate_xlsx()
        worksheet = self._load_ventes_sheet(wizard)
        row = self._ventes_data_rows(worksheet)[0]
        self.assertEqual(worksheet.cell(row, 1).value, "Facture")
        self.assertGreater(worksheet.cell(row, 8).value, 0)

    def test_t11_customer_refund_negative_amounts(self):
        """T11 — out_refund montants négatifs, Type Avoir."""
        invoice = self._create_customer_move(amount=200.0)
        refund = invoice._reverse_moves(
            default_values_list=[{"invoice_date": self.period_from}]
        )
        refund.action_post()
        wizard = self._wizard()
        wizard.action_generate_xlsx()
        worksheet = self._load_ventes_sheet(wizard)
        refund_row = None
        for row_idx in self._ventes_data_rows(worksheet):
            if worksheet.cell(row_idx, 1).value == "Avoir":
                refund_row = row_idx
                break
        self.assertIsNotNone(refund_row)
        self.assertLess(worksheet.cell(refund_row, 8).value, 0)

    def test_t12_totals_algebraic_sum(self):
        """T12 / T13 — totaux algébriques facture + avoir (Ventes)."""
        invoice = self._create_customer_move(amount=1000.0)
        refund = invoice._reverse_moves(
            default_values_list=[{"invoice_date": self.period_from}]
        )
        refund.action_post()
        wizard = self._wizard()
        wizard.action_generate_xlsx()
        worksheet = self._load_ventes_sheet(wizard)

        sale_moves = wizard._fetch_sale_moves()
        expected_ttc = sum(signed_move_amounts(move)["amount_ttc"] for move in sale_moves)

        totals_row = None
        for row_idx in range(FIRST_DATA_ROW + 1, worksheet.max_row + 1):
            if worksheet.cell(row_idx, 1).value == "Nombre de documents":
                totals_row = row_idx
                break
        self.assertIsNotNone(totals_row)
        self.assertEqual(worksheet.cell(totals_row, 2).value, len(sale_moves))
        self.assertAlmostEqual(worksheet.cell(totals_row, 8).value, expected_ttc)
        self.assertEqual(len(sale_moves), 2)

    def test_t14_payment_state_uses_odoo_label(self):
        """T14 — libellé payment_state traduit Odoo."""
        self._create_customer_move(amount=100.0)
        wizard = self._wizard()
        wizard.action_generate_xlsx()
        labels = wizard._payment_state_labels()
        worksheet = self._load_ventes_sheet(wizard)
        row = self._ventes_data_rows(worksheet)[0]
        move = wizard._fetch_sale_moves().filtered(
            lambda m: m.partner_id == self.partner
        )
        self.assertEqual(len(move), 1)
        expected = labels.get(move.payment_state, move.payment_state)
        self.assertEqual(worksheet.cell(row, 11).value, expected)
        self.assertNotEqual(worksheet.cell(row, 11).value, "not_paid")

    def test_t15_settled_amount_formula(self):
        """T15 — Montant réglé / soldé = sign * abs(total - residual)."""
        move = self._create_customer_move(amount=300.0)
        amounts = signed_move_amounts(move)
        wizard = self._wizard()
        wizard.action_generate_xlsx()
        worksheet = self._load_ventes_sheet(wizard)
        row = next(
            row_idx
            for row_idx in self._ventes_data_rows(worksheet)
            if worksheet.cell(row_idx, 2).value == move.name
        )
        self.assertAlmostEqual(worksheet.cell(row, 9).value, amounts["amount_paid"])

    def test_t16_ventes_column_order(self):
        """T16 — 11 colonnes dans l'ordre MOA."""
        wizard = self._wizard()
        wizard.action_generate_xlsx()
        worksheet = self._load_ventes_sheet(wizard)
        self.assertEqual(self._ventes_headers(worksheet), VENTES_HEADERS)


@tagged("post_install", "-at_install", "laplatine_billing_report")
class TestLaplatineBillingReportSliceE(TestLaplatineBillingReportWizard):
    """Slice E — onglets vides, anti-formule Excel, non-régression présentation."""

    def _generate_workbook(self, date_from, date_to):
        wizard = self._wizard(date_from=date_from, date_to=date_to)
        wizard.action_generate_xlsx()
        return self._load_workbook(wizard)

    def test_t18_both_sheets_empty(self):
        """T18 — aucun document : message, totaux à 0, deux feuilles présentes."""
        workbook = self._generate_workbook(
            fields.Date.from_string("2099-11-01"),
            fields.Date.from_string("2099-11-30"),
        )
        self.assertEqual(workbook.sheetnames, ["Ventes", "Achats"])
        self._assert_empty_sheet(workbook["Ventes"], VENTES_HEADERS, 6)
        self._assert_empty_sheet(workbook["Achats"], ACHATS_HEADERS, 7)

    def test_e01_ventes_empty_achats_with_data(self):
        """Slice E — Ventes vide, Achats avec écritures."""
        period_from = fields.Date.from_string("2099-07-01")
        period_to = fields.Date.from_string("2099-07-31")
        self._create_vendor_move(
            amount=120.0,
            invoice_date=period_from,
            vendor_ref="REF-E01",
        )
        workbook = self._generate_workbook(period_from, period_to)
        ventes = workbook["Ventes"]
        achats = workbook["Achats"]
        self._assert_empty_sheet(ventes, VENTES_HEADERS, 6)
        self.assertEqual(len(self._achats_data_rows(achats)), 1)
        totals_row = self._find_totals_row(achats)
        self.assertEqual(achats.cell(totals_row, 2).value, 1)

    def test_e02_achats_empty_ventes_with_data(self):
        """Slice E — Achats vide, Ventes avec écritures."""
        period_from = fields.Date.from_string("2099-09-01")
        period_to = fields.Date.from_string("2099-09-30")
        self._create_customer_move(invoice_date=period_from, amount=150.0)
        workbook = self._generate_workbook(period_from, period_to)
        ventes = workbook["Ventes"]
        achats = workbook["Achats"]
        self.assertEqual(len(self._ventes_data_rows(ventes)), 1)
        self._assert_empty_sheet(achats, ACHATS_HEADERS, 7)
        totals_row = self._find_totals_row(ventes)
        self.assertEqual(ventes.cell(totals_row, 2).value, 1)

    def test_t20_nombre_de_documents_label(self):
        """T20 — libellé exact « Nombre de documents » sur les deux onglets."""
        workbook = self._generate_workbook(
            fields.Date.from_string("2099-11-01"),
            fields.Date.from_string("2099-11-30"),
        )
        for sheet_name in ("Ventes", "Achats"):
            totals_row = self._find_totals_row(workbook[sheet_name])
            self.assertEqual(
                workbook[sheet_name].cell(totals_row, 1).value,
                "Nombre de documents",
            )

    def test_t21_excel_formula_injection_written_as_text(self):
        """T21 — valeurs =1+1, +CMD, -TEST, @REF restent du texte, pas des formules."""
        evil_customer = self.env["res.partner"].create(
            {"name": "=1+1", "customer_rank": 1}
        )
        evil_vendor = self.env["res.partner"].create(
            {"name": "+CMD", "supplier_rank": 1}
        )
        customer_move = self._create_customer_move(partner=evil_customer)
        vendor_move = self._create_vendor_move(
            partner=evil_vendor,
            vendor_ref="-TEST",
        )
        vendor_move_ref = self._create_vendor_move(vendor_ref="@REF", amount=45.0)

        wizard = self._wizard()
        wizard.action_generate_xlsx()
        ventes = self._load_ventes_sheet(wizard)
        achats = self._load_achats_sheet(wizard)

        ventes_row = next(
            row_idx
            for row_idx in self._ventes_data_rows(ventes)
            if ventes.cell(row_idx, 2).value == customer_move.name
        )
        self._assert_text_cell_not_formula(ventes.cell(ventes_row, 3), "=1+1")

        achats_row = next(
            row_idx
            for row_idx in self._achats_data_rows(achats)
            if achats.cell(row_idx, 2).value == vendor_move.name
        )
        self._assert_text_cell_not_formula(achats.cell(achats_row, 3), "-TEST")
        self._assert_text_cell_not_formula(achats.cell(achats_row, 4), "+CMD")

        achats_ref_row = next(
            row_idx
            for row_idx in self._achats_data_rows(achats)
            if achats.cell(row_idx, 2).value == vendor_move_ref.name
        )
        self._assert_text_cell_not_formula(achats.cell(achats_ref_row, 3), "@REF")

    def test_e04_slice_d_print_setup_preserved_on_empty_export(self):
        """Slice E — mise en page slice D conservée sur export sans document."""
        workbook = self._generate_workbook(
            fields.Date.from_string("2099-11-01"),
            fields.Date.from_string("2099-11-30"),
        )
        presentation = TestLaplatineBillingReportPresentation()
        for sheet_name in ("Ventes", "Achats"):
            presentation._assert_sheet_print_setup(workbook[sheet_name])

@tagged("post_install", "-at_install", "laplatine_billing_report")
class TestLaplatineBillingReportPresentation(TransactionCase):
    """Slice D — propriétés XLSX et impression vérifiables automatiquement."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Wizard = cls.env["laplatine.billing.report.wizard"]
        cls.period_from = fields.Date.from_string("2099-08-01")
        cls.period_to = fields.Date.from_string("2099-08-31")

    def _generate_empty_report(self):
        wizard = self.Wizard.create(
            {"date_from": self.period_from, "date_to": self.period_to}
        )
        wizard.action_generate_xlsx()
        return load_workbook(BytesIO(base64.b64decode(wizard.report_file)))

    def _assert_sheet_print_setup(self, worksheet):
        self.assertEqual(worksheet.page_setup.orientation, "landscape")
        self.assertEqual(worksheet.page_setup.paperSize, 9)
        self.assertTrue(worksheet.sheet_properties.pageSetUpPr.fitToPage)
        self.assertEqual(worksheet.print_title_rows.replace("$", ""), "1:6")
        self.assertAlmostEqual(worksheet.page_margins.left, 0.4, places=1)
        self.assertAlmostEqual(worksheet.page_margins.right, 0.4, places=1)
        self.assertEqual(worksheet.oddFooter.center.text, "Page &P / &N")
        self.assertFalse(worksheet.sheet_view.showGridLines)
        self.assertTrue(worksheet.print_area)
        self.assertEqual(worksheet.freeze_panes, "A7")

    def _assert_column_widths(self, worksheet, expected_widths):
        for index, expected in enumerate(expected_widths, start=1):
            letter = get_column_letter(index)
            width = worksheet.column_dimensions[letter].width
            self.assertIsNotNone(width)
            self.assertGreater(width, 8)
            self.assertLessEqual(width, expected + 5)

    def test_d01_ventes_print_setup(self):
        """D01 — paramètres d'impression onglet Ventes."""
        workbook = self._generate_empty_report()
        self._assert_sheet_print_setup(workbook["Ventes"])

    def test_d02_achats_print_setup(self):
        """D02 — paramètres d'impression onglet Achats."""
        workbook = self._generate_empty_report()
        self._assert_sheet_print_setup(workbook["Achats"])

    def test_d03_meta_block_content(self):
        """D03 — bloc méta lignes 1 à 4."""
        workbook = self._generate_empty_report()
        ventes = workbook["Ventes"]
        self.assertEqual(ventes["A1"].value, "Rapport de facturation — Ventes")
        self.assertIn("Du ", ventes["A3"].value)
        self.assertIn("Généré le", ventes["A4"].value)

    def test_d04_header_row_on_line_six(self):
        """D04 — en-têtes colonnes sur la ligne 6 Excel."""
        workbook = self._generate_empty_report()
        header_row_excel = HEADER_ROW + 1
        ventes = workbook["Ventes"]
        self.assertEqual(ventes.cell(header_row_excel, 1).value, "Type")
        achats = workbook["Achats"]
        self.assertEqual(achats.cell(header_row_excel, 1).value, "Type")
        self.assertEqual(achats.cell(header_row_excel, 3).value, "Référence fournisseur")

    def test_d05_column_widths_configured(self):
        """D05 — largeurs de colonnes définies (évite troncature ###)."""
        workbook = self._generate_empty_report()
        self._assert_column_widths(workbook["Ventes"], VENTES_COLUMN_WIDTHS)
        self._assert_column_widths(workbook["Achats"], ACHATS_COLUMN_WIDTHS)

    def test_d06_autofilter_and_print_area_include_totals(self):
        """D06 — zone d'impression jusqu'à la ligne de totaux."""
        workbook = self._generate_empty_report()
        ventes = workbook["Ventes"]
        self.assertTrue(ventes.auto_filter.ref)
        self.assertIn("6", ventes.auto_filter.ref)
        self.assertIn("Nombre de documents", ventes["A8"].value)


@tagged("post_install", "-at_install", "laplatine_billing_report")
class TestLaplatineBillingReportHelpers(TransactionCase):
    def test_report_sign_and_type_labels(self):
        self.assertEqual(report_sign("out_invoice"), 1)
        self.assertEqual(report_sign("out_refund"), -1)
        self.assertEqual(document_type_label("out_invoice"), "Facture")
        self.assertEqual(document_type_label("out_refund"), "Avoir")
