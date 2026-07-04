# -*- coding: utf-8 -*-
from io import BytesIO
from unittest.mock import Mock

from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from odoo import fields
from odoo.exceptions import UserError
from odoo.tests import TransactionCase, tagged

from ..report.customer_statement_xlsx import (
    FIRST_DATA_ROW,
    HEADER_ROW,
    LaplatineCustomerStatementXlsx,
    compute_report_summary,
    invoice_display_status,
    is_due_date_overdue,
    status_uses_overdue_highlight,
)


def _invoice_stub(payment_state, amount_total, amount_residual, invoice_date_due=None):
    return Mock(
        payment_state=payment_state,
        amount_total=amount_total,
        amount_residual=amount_residual,
        invoice_date_due=invoice_date_due,
    )


@tagged("post_install", "-at_install", "laplatine_customer_statement")
class TestLaplatineCustomerStatementStatus(TransactionCase):
    def test_paid_invoice_status(self):
        ref = fields.Date.from_string("2026-07-04")
        status = invoice_display_status("paid", 100.0, 0.0, "2026-06-01", ref)
        self.assertEqual(status, "Payée")

    def test_unpaid_future_due_status(self):
        ref = fields.Date.from_string("2026-07-04")
        status = invoice_display_status("not_paid", 100.0, 100.0, "2026-07-10", ref)
        self.assertEqual(status, "À payer")

    def test_unpaid_overdue_status(self):
        ref = fields.Date.from_string("2026-07-04")
        status = invoice_display_status("not_paid", 100.0, 100.0, "2026-07-03", ref)
        self.assertEqual(status, "En retard")

    def test_partial_not_overdue_status(self):
        ref = fields.Date.from_string("2026-07-04")
        status = invoice_display_status("partial", 100.0, 40.0, "2026-07-04", ref)
        self.assertEqual(status, "Partiellement payée")

    def test_partial_overdue_status(self):
        ref = fields.Date.from_string("2026-07-04")
        status = invoice_display_status("partial", 100.0, 40.0, "2026-07-01", ref)
        self.assertEqual(status, "Partiellement payée — en retard")

    def test_due_date_equal_to_today_not_overdue(self):
        ref = fields.Date.from_string("2026-07-04")
        self.assertFalse(is_due_date_overdue(ref, ref))
        status = invoice_display_status("not_paid", 50.0, 50.0, ref, ref)
        self.assertEqual(status, "À payer")

    def test_in_payment_status(self):
        ref = fields.Date.from_string("2026-07-04")
        status = invoice_display_status("in_payment", 100.0, 100.0, "2026-06-01", ref)
        self.assertEqual(status, "Paiement en cours")

    def test_missing_due_date_not_marked_overdue(self):
        ref = fields.Date.from_string("2026-07-04")
        status = invoice_display_status("not_paid", 100.0, 100.0, False, ref)
        self.assertEqual(status, "À payer")

    def test_summary_total_to_pay_and_overdue(self):
        ref = fields.Date.from_string("2026-07-04")
        invoices = [
            _invoice_stub("paid", 100.0, 0.0, "2026-06-01"),
            _invoice_stub("not_paid", 200.0, 200.0, "2026-07-01"),
            _invoice_stub("partial", 150.0, 50.0, "2026-07-10"),
        ]
        summary = compute_report_summary(invoices, ref)
        self.assertEqual(summary["total_invoiced"], 450.0)
        self.assertEqual(summary["total_paid"], 200.0)
        self.assertEqual(summary["total_to_pay"], 250.0)
        self.assertEqual(summary["total_overdue"], 200.0)

    def test_overdue_highlight_detection(self):
        self.assertTrue(status_uses_overdue_highlight("En retard"))
        self.assertTrue(status_uses_overdue_highlight("Partiellement payée — en retard"))
        self.assertFalse(status_uses_overdue_highlight("À payer"))


@tagged("post_install", "-at_install", "laplatine_customer_statement")
class TestLaplatineCustomerStatement(TransactionCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.company = cls.env.company
        cls.partner = cls.env["res.partner"].create(
            {
                "name": "Client Test La Platine Statement",
                "customer_rank": 1,
            }
        )
        cls.child_partner = cls.env["res.partner"].create(
            {
                "name": "Adresse facturation test",
                "type": "invoice",
                "parent_id": cls.partner.id,
            }
        )
        cls.sale_journal = cls.env["account.journal"].search(
            [("type", "=", "sale")],
            limit=1,
        )
        cls.income_account = cls.env["account.account"].search(
            [("account_type", "=", "income")],
            limit=1,
        )
        assert cls.sale_journal, "Journal de vente requis pour les tests."
        assert cls.income_account, "Compte de produit requis pour les tests."

    def _wizard(self, partner=None, date_from=None, date_to=None):
        if date_from is None or date_to is None:
            date_from, date_to = self.env[
                "laplatine.customer.statement.wizard"
            ]._default_period_bounds()
        return self.env["laplatine.customer.statement.wizard"].create(
            {
                "partner_id": (partner or self.partner).id,
                "date_from": date_from,
                "date_to": date_to,
            }
        )

    def _create_posted_invoice(
        self,
        partner=None,
        amount=100.0,
        invoice_date=None,
        due_date=None,
        currency=None,
    ):
        partner = partner or self.partner
        if invoice_date is None:
            invoice_date = self._wizard().date_from
        values = {
            "move_type": "out_invoice",
            "partner_id": partner.id,
            "invoice_date": invoice_date,
            "invoice_date_due": due_date,
            "journal_id": self.sale_journal.id,
            "invoice_line_ids": [
                (
                    0,
                    0,
                    {
                        "name": "Ligne test relevé client",
                        "quantity": 1,
                        "price_unit": amount,
                        "account_id": self.income_account.id,
                    },
                )
            ],
        }
        if currency:
            values["currency_id"] = currency.id
        move = self.env["account.move"].create(values)
        move.action_post()
        return move

    def test_default_period_is_last_90_days_including_today(self):
        wizard = self._wizard()
        today = fields.Date.today()
        expected_from = today - relativedelta(days=89)
        self.assertEqual(wizard.date_from, expected_from)
        self.assertEqual(wizard.date_to, today)

    def test_domain_includes_child_partner_invoices(self):
        invoice = self._create_posted_invoice(partner=self.child_partner)
        wizard = self._wizard(partner=self.partner)
        invoices = wizard._fetch_invoices()
        self.assertIn(invoice, invoices)

    def test_reversed_invoice_excluded(self):
        invoice = self._create_posted_invoice(amount=50.0)
        reversal = invoice._reverse_moves(
            default_values_list=[{"invoice_date": invoice.invoice_date}]
        )
        reversal.action_post()
        wizard = self._wizard()
        invoices = wizard._fetch_invoices()
        self.assertNotIn(invoice, invoices)

    def test_no_invoice_raises_user_error(self):
        wizard = self._wizard()
        with self.assertRaises(UserError):
            wizard.action_generate_xlsx()

    def test_multi_currency_blocked(self):
        usd = self.env["res.currency"].search([("name", "=", "USD")], limit=1)
        self.assertTrue(usd)
        self._create_posted_invoice(amount=100.0)
        self._create_posted_invoice(amount=80.0, currency=usd)
        wizard = self._wizard()
        with self.assertRaises(UserError):
            wizard.action_generate_xlsx()

    def test_xlsx_workbook_properties_summary_and_totals(self):
        today = fields.Date.today()
        invoice_a = self._create_posted_invoice(
            amount=100.0,
            invoice_date=today,
            due_date=today + relativedelta(days=10),
        )
        invoice_b = self._create_posted_invoice(
            amount=40.0,
            invoice_date=today,
            due_date=today - relativedelta(days=5),
        )
        wizard = self._wizard(date_from=today - relativedelta(days=30), date_to=today)
        invoices = wizard._fetch_invoices()
        generator = LaplatineCustomerStatementXlsx(
            partner=self.partner,
            date_from=wizard.date_from,
            date_to=wizard.date_to,
            invoices=invoices,
            generation_date=today,
        )
        content, filename = generator.generate()

        self.assertTrue(filename.endswith(".xlsx"))
        self.assertIn("Etat_facturation_", filename)

        workbook = load_workbook(BytesIO(content))
        worksheet = workbook.active
        header_row_excel = HEADER_ROW + 1
        first_data_row_excel = FIRST_DATA_ROW + 1
        totals_row_excel = first_data_row_excel + len(invoices)

        self.assertEqual(worksheet.title, "État de facturation")
        self.assertEqual(worksheet["A1"].value, "État de facturation")
        self.assertEqual(worksheet.cell(header_row_excel, 1).value, "Facture")
        self.assertEqual(worksheet["A5"].value, "Total facturé")
        self.assertEqual(
            worksheet["A7"].value,
            "Montant total à régler à La Platine",
        )
        self.assertEqual(worksheet["A8"].value, "Dont montant en retard")

        self.assertEqual(worksheet.page_setup.orientation, "landscape")
        self.assertEqual(worksheet.page_setup.paperSize, 9)
        self.assertTrue(worksheet.sheet_properties.pageSetUpPr.fitToPage)
        self.assertIn(
            worksheet.print_title_rows.replace("$", ""),
            (f"{header_row_excel}:{header_row_excel}",),
        )
        self.assertAlmostEqual(worksheet.page_margins.left, 0.4)
        self.assertEqual(worksheet.oddFooter.center.text, "Page &P / &N")
        self.assertFalse(worksheet.sheet_view.showGridLines)

        summary = compute_report_summary(invoices, today, currency=invoices[:1].currency_id)
        self.assertEqual(worksheet["D5"].value, summary["total_invoiced"])
        self.assertEqual(worksheet["D7"].value, summary["total_to_pay"])
        self.assertEqual(worksheet["D8"].value, summary["total_overdue"])

        totals_row = list(
            worksheet.iter_rows(
                min_row=totals_row_excel,
                max_row=totals_row_excel,
                min_col=1,
                max_col=7,
                values_only=True,
            )
        )[0]
        self.assertEqual(totals_row[2], "Totaux")
        self.assertEqual(totals_row[3], invoice_a.amount_total + invoice_b.amount_total)
        self.assertEqual(
            totals_row[5],
            invoice_a.amount_residual + invoice_b.amount_residual,
        )
